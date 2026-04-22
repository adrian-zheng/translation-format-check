from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("output/spreadsheet")

ISSUE_SENTENCES = [
    '“The Aurora desk”—designed for hybrid teams…  brings  focus to shared work.',
    "The supplier–not the distributor—must confirm “final acceptance”… before shipment.",
    "Customers said ‘it feels premium’……  but the launch copy still  needs cleanup.",
    "Module 3—Quality Review explains how to flag “ambiguous terms”  and  revise them.",
    "  Leading spaces should disappear—and trailing spaces should too.  ",
    "The button label–if translated literally—may confuse users… please review.",
    "“Confirm”—then proceed to the next step…  Do not close the tab.",
    "This clean row should stay unchanged.",
]

SOURCE_SENTENCES = [
    "这是一条用于测试的产品文案。",
    "供应商需要在发货前确认最终验收。",
    "客户反馈语气需要更自然。",
    "培训材料需要检查术语一致性。",
    "这里用于测试行首和行尾空格。",
    "按钮文案需要确认是否清晰。",
    "操作提示需要符合英文格式。",
    "这是一条无需修改的干净记录。",
]


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    write_csv(OUTPUT_DIR / "translation_format_small.csv", 16)
    write_csv(OUTPUT_DIR / "translation_format_100_rows.csv", 100)
    write_xlsx(OUTPUT_DIR / "translation_format_small.xlsx", 20)
    write_xlsx(OUTPUT_DIR / "translation_format_2000_rows.xlsx", 2000)


def make_row(index: int) -> dict[str, str]:
    sentence_index = (index - 1) % len(ISSUE_SENTENCES)
    return {
        "ID": f"T-{index:04d}",
        "中文原文": SOURCE_SENTENCES[sentence_index],
        "英文译文": ISSUE_SENTENCES[sentence_index],
        "Reviewer": ["Alex", "Morgan", "Riley", "Casey"][index % 4],
        "备注": "fake data for local testing",
    }


def write_csv(path: Path, count: int) -> None:
    fieldnames = ["ID", "中文原文", "英文译文", "Reviewer", "备注"]
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for index in range(1, count + 1):
            writer.writerow(make_row(index))


def write_xlsx(path: Path, count: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Translation QA"
    headers = ["ID", "中文原文", "English Translation", "Reviewer", "备注"]
    ws.append(headers)

    for index in range(1, count + 1):
        row = make_row(index)
        ws.append([
            row["ID"],
            row["中文原文"],
            row["英文译文"],
            row["Reviewer"],
            row["备注"],
        ])

    style_sheet(ws)
    wb.save(path)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="0A6E69")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    widths = [14, 34, 72, 16, 28]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(wrap_text=True, vertical="top")


if __name__ == "__main__":
    main()
